import re
import csv
from print_telefone import PrintTelefone
from investimento import Investimento
import openpyxl
from openpyxl.styles import Alignment, Font


def extract_currency(value):
    cleaned_value = re.sub(r'[^\d,]', '', value)
    cleaned_value = cleaned_value.replace(',', '.')
    return float(cleaned_value)

def process_texto_bruto(texto_bruto):
    linhas = texto_bruto.splitlines()
    print_telefone = PrintTelefone()
    investimento_atual = None
    
    for line in linhas:
        cleaned_line = line.strip()
        
        if cleaned_line.startswith("RS$") or cleaned_line.startswith("RS"):
            print_telefone.investimento_liquido = extract_currency(cleaned_line)
            continue
        
        sigla_match = re.search(r'\b[A-Z]{3}\b', cleaned_line)
        porcentagem_match = re.search(r'(\d+[\,\.]?\d*)%', cleaned_line)
        valor_match = re.search(r'RS?\$?(\d{1,3}(?:\.\d{3})*,\d{2})', cleaned_line)

        if sigla_match:
            sigla = sigla_match.group()
            investimento_atual = Investimento(sigla=sigla, porcentagem=0.0, valor=0.0)
            print_telefone.investimentos.append(investimento_atual)
        
        if investimento_atual:
            if porcentagem_match:
                porcentagem_str = porcentagem_match.group(1).replace('%', '')
                investimento_atual.porcentagem = float(porcentagem_str.replace(",", "."))
            if valor_match:
                investimento_atual.valor = extract_currency(valor_match.group(1))

    total_investimentos = sum(investimento.valor for investimento in print_telefone.investimentos)
    if total_investimentos > print_telefone.investimento_liquido:
        print_telefone.investimento_liquido = total_investimentos

    return print_telefone

def save_to_csv(extracted_text, output_path="output.csv"):
    print_telefone = process_texto_bruto(extracted_text)

    with open(output_path, mode="w", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=';')  # ponto e vírgula é o delimitador padrão de um arquivo .csv
        
        writer.writerow(["XP Ável", "Valor total da carteira"])
        writer.writerow(["", f"R$ {print_telefone.investimento_liquido:.2f}"])
        
        writer.writerow([])
        writer.writerow(["Classe Ativo", "Nome Ativo", "Valor Total", "% da carteira"])

        for investimento in print_telefone.investimentos:
            writer.writerow([
                "",  # Coluna para Classe Ativo
                investimento.sigla,
                f"R$ {investimento.valor:.2f}",
                f"{investimento.porcentagem:.2f}%"
            ])
    
    print(f"Arquivo salvo com sucesso em {output_path}")

def save_to_excel(extracted_text, output_path="output.xlsx"):
    print_telefone = process_texto_bruto(extracted_text)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Investimentos"

    bold_font = Font(bold=True)

    # Título "XP Ável" e "Valor total da carteira" que se estendem sobre as colunas
    sheet.merge_cells("A1:B1")
    sheet.merge_cells("C1:D1")
    sheet["A1"] = "XP Ável"
    sheet["C1"] = "Valor total da carteira"
    
    # Valor do investimento líquido que se estende sobre as colunas "Valor Total" e "% da carteira"
    sheet.merge_cells("C2:D2")
    sheet["C2"] = f"{print_telefone.investimento_liquido:.2f}"

    sheet["A1"].font = bold_font
    sheet["C1"].font = bold_font
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet["C1"].alignment = Alignment(horizontal="center")
    sheet["C2"].alignment = Alignment(horizontal="center")

    headers = ["Classe Ativo", "Nome Ativo", "Valor Total", "% da carteira"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    for idx, investimento in enumerate(print_telefone.investimentos, start=4):
        sheet.cell(row=idx, column=1, value="")  # Classe Ativo
        sheet.cell(row=idx, column=2, value=investimento.sigla)  # Nome Ativo
        sheet.cell(row=idx, column=3, value=f"{investimento.valor:.2f}")  # Valor Total
        sheet.cell(row=idx, column=4, value=f"{investimento.porcentagem:.2f}%")  # % da carteira

    column_widths = [15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    workbook.save(output_path)
    print(f"Arquivo salvo em {output_path}")
