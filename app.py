import tempfile
import os
from preprocess import preprocess_image
from status_bar import auto_detect_and_remove_bars
from ocr import extract_text_from_image
from save_to_file import process_texto_bruto
import streamlit as st
from print_telefone import PrintTelefone
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def gerar_excel_formatado(dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Investimentos"

    bold_font = Font(bold=True)
    
    ws.merge_cells("A1:B1")
    ws.merge_cells("C1:D1")
    ws["A1"] = "XP Ável"
    ws["C1"] = "Valor total da carteira"

    ws.merge_cells("C2:D2")
    ws["C2"] = f"{dados.investimento_liquido:.2f}"

    ws["A1"].font = bold_font
    ws["C1"].font = bold_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["C1"].alignment = Alignment(horizontal="center")
    ws["C2"].alignment = Alignment(horizontal="center")

    headers = ["Classe Ativo", "Nome Ativo", "Valor Total", "% da carteira"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    for idx, investimento in enumerate(dados.investimentos, start=4):
        ws.cell(row=idx, column=1, value="")  # Classe Ativo
        ws.cell(row=idx, column=2, value=investimento.sigla)  # Nome Ativo
        ws.cell(row=idx, column=3, value=f"{investimento.valor:.2f}")  # Valor Total
        ws.cell(row=idx, column=4, value=f"{investimento.porcentagem:.2f}%")  # % da carteira

    column_widths = [15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return tmp.name


def processa_st(dados: PrintTelefone):
    if st.button("Processar Imagem"):
        excel_path = gerar_excel_formatado(dados)
        
        st.subheader("Valor Total da Carteira:")
        st.write(f"R$ {dados.investimento_liquido:.2f}")
        
        st.subheader("Planilha processada com :green[sucesso] :white_check_mark:")

        with open(excel_path, "rb") as file:
            st.download_button(
                label="Baixar Planilha",
                data=file,
                file_name="planilha_investimentos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        
        os.remove(excel_path)
        
def main_pipeline(image_path):
    processed_image_path = preprocess_image(image_path)

    processed_image_path = auto_detect_and_remove_bars(processed_image_path)

    extracted_text: str = extract_text_from_image(processed_image_path)
    
    dados: PrintTelefone = process_texto_bruto(extracted_text)
    
    processa_st(dados)

if __name__ == "__main__":
    st.title("Portfácil")

    uploaded_file = st.file_uploader("Escolha uma imagem", type=["jpeg", "jpg", "png"])
    
    if uploaded_file is not None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
            tmp.write(uploaded_file.getvalue())
            image_path = tmp.name
        
        main_pipeline(image_path)
